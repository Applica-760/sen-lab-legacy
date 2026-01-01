"""
画像ファイル名からシート名・時刻（行）を導出し、該当セル(C列)を起点として
下へ300行ぶんのC列値を取得 → 指定規則で {0,1,2} に変換 → 300桁を連結文字列化。
さらに「全300値が同一なら1、異なれば0」のフラグを付け、CSVに
[ file_path, concatenated_300, uniform_flag ] の3列で書き出します。

前提（既存仕様を踏襲）:
- 画像ファイル名は例: '08_000122_ffmpeg.jpg'
  - 先頭2桁 = 日付 (DD)
  - 次の6桁 = 時刻 (HHMMSS)
  - "000122" は 00:01:22 を表す
- Excelの各シートは '3_05DD' の形式（例: 3_0508）
  - 月は --month で指定（デフォルト '05'）
  - シート接頭辞は --sheet-prefix で指定（デフォルト '3_'）
- Excelの A2=0:00:00, A3=0:00:01, … の対応なので、
  HH:MM:SS を秒に変換した値 + 2 が行番号（1-indexed）になります。
- C列を起点に、下方向へ 300 行ぶん取得します（起点を含む）。

値の変換規則:
  期待する値は整数。
  0, 1, 4以上 → 0
  2           → 1
  3           → 2
  非数/欠損などは 0 とみなします。

使い方例:
python tools/dataprep/get_300seqs.py \
    --excel data/teaching_data_fully.xlsx \
    --images-root data/all_s \
    --output data/get_300seqs.csv \
    --jobs 10
    
"""

import argparse
from PIL import Image
import csv
import re
from pathlib import Path
from typing import Dict, List, Tuple, Optional
from concurrent.futures import ProcessPoolExecutor, as_completed
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# --------- 変換ロジック ---------
def coerce_to_int(value) -> int:
    if value is None:
        return 0
    try:
        return int(round(float(value)))
    except Exception:
        return 0


def transform_value(iv: int) -> int:
    # 0,1,4以上→0 / 2→1 / 3→2
    if iv == 2:
        return 1
    if iv == 3:
        return 2
    return 0


def hhmmss_to_seconds(hms: str) -> int:
    if len(hms) != 6 or not hms.isdigit():
        raise ValueError(f"Invalid HHMMSS: {hms}")
    hh = int(hms[0:2])
    mm = int(hms[2:4])
    ss = int(hms[4:6])
    return hh * 3600 + mm * 60 + ss


# --------- 画像名→(シート, 行) 変換 ---------
def filename_to_sheet_and_row(filename: str, pat: re.Pattern, sheet_prefix: str, month: str) -> Tuple[str, int]:
    m = pat.fullmatch(filename)
    if not m:
        raise ValueError(f"Filename does not match pattern: {filename}")
    day = m.group("day")      # '08'
    hms = m.group("hms")      # '000122'
    row_index = hhmmss_to_seconds(hms) + 2   # A2=0秒 → 行=秒+2
    sheet_name = f"{sheet_prefix}{month}{day}"
    return sheet_name, row_index


# --------- C列の一括抽出 ---------
def load_column_c_as_array(ws: Worksheet) -> List[int]:
    max_row = ws.max_row
    values = [0] * (max_row + 1)
    for i, (v,) in enumerate(ws.iter_rows(min_col=3, max_col=3, min_row=1, max_row=max_row, values_only=True), start=1):
        values[i] = coerce_to_int(v)
    return values


# --------- 画像の横幅取得 ---------
def get_image_width(path: Path) -> int:
    """画像の横幅を取得。読み取れない場合は -1 を返す。"""
    try:
        with Image.open(path) as im:
            return int(im.size[0])
    except Exception:
        return -1


# --------- 1シート分処理 ---------
def process_tasks_for_sheet(
    sheet_name: str,
    tasks: List[Tuple[Path, int]],
    c_values: List[int],
    window: int
) -> List[Tuple[str, str, int, int]]:
    max_row = len(c_values) - 1
    out_rows: List[Tuple[str, str, int, int]] = []

    for img_path, start_row in tasks:
        end_row = start_row + window - 1
        if start_row <= max_row:
            last = min(end_row, max_row)
            slice_vals = c_values[start_row:last + 1]
            if len(slice_vals) < window:
                slice_vals += [0] * (window - len(slice_vals))
        else:
            slice_vals = [0] * window

        mapped = [transform_value(v) for v in slice_vals]
        uniform_flag = 1 if (min(mapped) == max(mapped)) else 0
        concatenated = "".join("012"[m] if m in (0, 1, 2) else "0" for m in mapped)
        width = get_image_width(img_path)
        out_rows.append((str(img_path), concatenated, uniform_flag, width))

    return out_rows


# --------- 並列ワーカ ---------
def _sheet_worker(args):
    sheet_name, tasks, c_values, window = args
    return sheet_name, process_tasks_for_sheet(sheet_name, tasks, c_values, window)


# --------- ユーティリティ ---------
def iter_image_files(root: Path, exts: List[str]) -> List[Path]:
    exts = {e.lower() for e in exts}
    return [p for p in root.rglob("*") if p.is_file() and p.suffix.lower() in exts]


# --------- メイン ---------
def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="C列300行を連結・変換してCSV出力するツール（画像幅付き）")
    p.add_argument("--excel", required=True)
    p.add_argument("--images-root", required=True)
    p.add_argument("--output", required=True)
    p.add_argument("--month", default="05")
    p.add_argument("--sheet-prefix", default="3_")
    p.add_argument("--pattern", default=r"(?P<day>\d{2})_(?P<hms>\d{6})_ffmpeg\.jpg")
    p.add_argument("--exts", nargs="+", default=[".jpg", ".jpeg", ".png"])
    p.add_argument("--window", type=int, default=300)
    p.add_argument("--jobs", type=int, default=1)
    p.add_argument("--log-every", type=int, default=0)
    return p.parse_args()


def main():
    args = parse_args()
    excel_path = Path(args.excel).expanduser().resolve()
    images_root = Path(args.images_root).expanduser().resolve()
    out_csv = Path(args.output).expanduser().resolve()

    if not excel_path.exists():
        raise FileNotFoundError(f"Excel not found: {excel_path}")
    if not images_root.exists():
        raise FileNotFoundError(f"Images root not found: {images_root}")

    files = iter_image_files(images_root, args.exts)
    if not files:
        print("画像が見つかりません。")
        return

    pat = re.compile(args.pattern)
    grouped: Dict[str, List[Tuple[Path, int]]] = {}
    skipped = 0
    for p in sorted(files):
        try:
            sheet, row = filename_to_sheet_and_row(p.name, pat, args.sheet_prefix, args.month)
            grouped.setdefault(sheet, []).append((p, row))
        except Exception:
            skipped += 1

    wb = load_workbook(filename=str(excel_path), data_only=True, read_only=True)
    grouped = {s: t for s, t in grouped.items() if s in wb.sheetnames}

    sheet_c_arrays = {s: load_column_c_as_array(wb[s]) for s in grouped.keys()}

    results: List[Tuple[str, str, int, int]] = []
    for s, tasks in grouped.items():
        results.extend(process_tasks_for_sheet(s, tasks, sheet_c_arrays[s], args.window))

    out_csv.parent.mkdir(parents=True, exist_ok=True)
    with out_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["file_path", "converted_300", "uniform_flag", "image_width_px"])
        w.writerows(results)

    print(f"完了: 出力 {len(results)} 件, スキップ {skipped} 件 → {out_csv}")


if __name__ == "__main__":
    main()